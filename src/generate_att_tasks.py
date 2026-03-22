import openpyxl
import json
import re
import hashlib
from datetime import datetime, timedelta
import random

random.seed(42)

wb = openpyxl.load_workbook(r'C:\CursorData\ObsAgents\src\att.xlsx')
ws = wb.active

rows = []
for r in range(2, ws.max_row + 1):
    wf = ws.cell(r, 1).value
    if not wf:
        continue
    rows.append({
        'workflow': str(wf).strip(),
        'description': str(ws.cell(r, 2).value or '').strip(),
        'exception': str(ws.cell(r, 3).value or '').strip(),
        'error_code': str(ws.cell(r, 4).value or '').strip(),
        'classification': str(ws.cell(r, 5).value or '').strip(),
        'resolution': str(ws.cell(r, 6).value or '').strip(),
        'task_step': str(ws.cell(r, 7).value or '').strip(),
        'task_type': str(ws.cell(r, 8).value or '').strip(),
        'resolution_type_desc': str(ws.cell(r, 9).value or '').strip(),
        'resolution_name': str(ws.cell(r, 10).value or '').strip(),
        'more_info': str(ws.cell(r, 11).value or '').strip(),
    })

wf_groups = {}
for row in rows:
    wf = row['workflow']
    if wf not in wf_groups:
        wf_groups[wf] = []
    wf_groups[wf].append(row)

TYPE_MAP = {
    'Hybrid': 'HYBRID',
    'Automated': 'AUTO-AGENT',
    'Manual': 'MANUAL',
}

PRIORITIES = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']
ASSIGNEES = [
    'Michael Torres', 'Rachel Kim', 'James Wright', 'Sarah Martinez',
    'David Cohen', 'Emily Zhang', 'Robert Patel', 'Lisa Anderson',
    'Chris Johnson', 'Anna Kovalenko', 'Mark Reeves', 'Diana Ruiz',
    'Brian Thompson', 'Priya Sharma', 'Kevin Oconnell', 'Julia Barnes',
]

MORE_INFO_CONDITION = (
    "If error item needs to be dropped from RTB, click on DROP button. "
    "If issue is already resolved by operations, click Complete Without Activity."
)

def clean_text(s):
    if not s or s == 'None':
        return ''
    s = s.replace('\u00e2\u20ac\u201c', '-').replace('\u00e2\u20ac\u201d', '-')
    s = s.replace('\uf0b7', '*')
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
    s = re.sub(r'[^\x20-\x7e\n\r\t]', '', s)
    return s.strip()

def detect_conditions(steps_data):
    """Detect conditional branches: steps with 'Case N:' or 'When ...:' in resolution."""
    case_pattern = re.compile(r'(?:Case\s+\d+|When\s+[\w\-\.]+)', re.IGNORECASE)
    condition_groups = {}
    for i, s in enumerate(steps_data):
        res = s['resolution_name']
        match = case_pattern.search(res)
        if match:
            base_label = s['task_step']
            if base_label not in condition_groups:
                condition_groups[base_label] = []
            condition_groups[base_label].append({
                'index': i,
                'condition': match.group(0).strip(),
            })
    return condition_groups

def has_rtb_condition(steps_data):
    """Check if workflow has the RTB drop/complete condition pattern."""
    for s in steps_data:
        mi = s.get('more_info', '').lower()
        if 'drop' in mi and 'rtb' in mi:
            return True
    return False

tasks = []
wf_list = sorted(wf_groups.keys())

for wf_idx, wf_name in enumerate(wf_list):
    steps_data = wf_groups[wf_name]
    task_id = f"ATT-{10001 + wf_idx}"
    first = steps_data[0]

    priority_hash = int(hashlib.md5(wf_name.encode()).hexdigest()[:8], 16)
    priority = PRIORITIES[priority_hash % 4]

    base_date = datetime(2026, 3, 8, 8, 0)
    created_offset = timedelta(hours=wf_idx * 0.3)
    created = base_date + created_offset
    sla_offset = timedelta(hours=random.randint(12, 96))

    assignee = ASSIGNEES[wf_idx % len(ASSIGNEES)]
    is_agent = any(s['task_type'] == 'Automated' for s in steps_data)

    summary_desc = clean_text(first['description'])
    if not summary_desc:
        summary_desc = f"{wf_name} workflow fallout"

    exception_text = clean_text(first['exception'])
    classification_text = clean_text(first['classification'])
    error_code = clean_text(first['error_code'])

    error_snippet_parts = {}
    if classification_text and classification_text != 'NA':
        error_snippet_parts['classification'] = classification_text[:80]
    if error_code and error_code != 'NA':
        error_snippet_parts['error_code'] = error_code[:60]
    if exception_text and exception_text != 'NA':
        error_snippet_parts['exception'] = exception_text[:120]
    error_snippet_parts['workflow'] = wf_name
    error_snippet_parts['msg'] = clean_text(first.get('resolution', ''))[:80] or 'Processing error'
    error_snippet = json.dumps(error_snippet_parts, indent=2)

    rtb_condition = has_rtb_condition(steps_data)
    conditions = detect_conditions(steps_data)

    total_steps = len(steps_data)
    current_step_raw = min(2, total_steps)
    if total_steps <= 2:
        current_step_raw = 1

    steps = []
    step_id_base = (wf_idx + 1) * 100
    condition_parent_map = {}

    for si, sd in enumerate(steps_data):
        step_id = step_id_base + si + 1
        step_label = clean_text(sd['task_step']) or f"Step {si+1}"
        step_type = TYPE_MAP.get(sd['task_type'], 'MANUAL')
        desc = clean_text(sd['resolution_type_desc'])
        resolution = clean_text(sd['resolution_name'])

        step_obj = {
            'id': step_id,
            'type': step_type,
            'label': step_label,
        }
        if desc:
            step_obj['description'] = desc
        if resolution:
            step_obj['resolution'] = resolution

        case_match = re.search(r'(?:Case\s+(\d+)|When\s+([\w\-\.]+))', resolution, re.IGNORECASE)
        if case_match:
            cond_label = case_match.group(0).strip()
            step_obj['conditionLabel'] = cond_label

            if step_label in condition_parent_map:
                parent_id = condition_parent_map[step_label]
                step_obj['conditionParent'] = parent_id
            else:
                condition_parent_map[step_label] = step_id

        steps.append(step_obj)

    if rtb_condition and len(steps) >= 2:
        first_investigate = None
        for s in steps:
            if 'Check' in s['label'] or 'Verify' in s['label'] or 'Investigate' in s['label']:
                first_investigate = s['id']
                break
        if first_investigate:
            has_drop = any('drop' in s['label'].lower() or 'rtb' in s.get('resolution', '').lower()
                          for s in steps if s['id'] != first_investigate)
            has_complete = any('complete' in s['label'].lower() or 'close' in s['label'].lower()
                              for s in steps if s['id'] != first_investigate)
            if not has_drop and not has_complete:
                drop_id = step_id_base + len(steps) + 1
                complete_id = step_id_base + len(steps) + 2
                steps.append({
                    'id': drop_id,
                    'type': 'MANUAL',
                    'label': 'Drop from RTB',
                    'description': 'Drop the error item from RTB and add reason code.',
                    'resolution': 'Click DROP button and add reason code.',
                    'conditionLabel': 'If drop needed',
                    'conditionParent': first_investigate,
                })
                steps.append({
                    'id': complete_id,
                    'type': 'AUTO-AGENT',
                    'label': 'Complete without activity',
                    'description': 'Close the task if issue is already resolved by operations team.',
                    'resolution': 'Click Complete Without Activity to close the task.',
                    'conditionLabel': 'If already resolved',
                    'conditionParent': first_investigate,
                })

    for si, step in enumerate(steps):
        if si < current_step_raw - 1:
            step['status'] = 'DONE'
        elif si == current_step_raw - 1:
            step['status'] = 'IN_PROGRESS'
        else:
            step['status'] = 'PENDING'

    if steps and current_step_raw >= 2:
        first_step = steps[0]
        if first_step['type'] == 'HYBRID' and 'agentLog' not in first_step:
            first_step['agentLog'] = [
                f"[{created.strftime('%H:%M')}:01] Initiating {first_step['label']}...",
                f"[{created.strftime('%H:%M')}:03] Analyzing error context for {wf_name}...",
                f"[{created.strftime('%H:%M')}:04] Check completed. Proceeding to next step.",
            ]

    history = [
        {'date': created.strftime('%b %d, %H:%M'), 'text': 'Ticket created by System'},
        {'date': (created + timedelta(minutes=2)).strftime('%b %d, %H:%M'), 'text': f'Assigned to {assignee}'},
    ]
    if current_step_raw >= 2:
        history.append({
            'date': (created + timedelta(minutes=5)).strftime('%b %d, %H:%M'),
            'text': f'Step 1 completed',
            'active': True,
        })
    else:
        history[-1]['active'] = True

    wf_desc = clean_text(first['description'])
    if len(wf_desc) > 150:
        wf_desc = wf_desc[:147] + '...'
    if not wf_desc:
        wf_desc = f"Resolve {wf_name} fallout and restore normal operation."

    obs_parts = []
    if classification_text and classification_text != 'NA':
        obs_parts.append(f"Classification: {classification_text[:100]}.")
    if exception_text and exception_text != 'NA':
        obs_parts.append(f"Exception pattern: {exception_text[:100]}.")
    if wf_desc:
        obs_parts.append(f"Context: {wf_desc[:120]}.")
    if not obs_parts:
        obs_parts.append(f"Workflow {wf_name} triggered for investigation.")
    observability = ' '.join(obs_parts)

    task = {
        'id': task_id,
        'priority': priority,
        'summary': summary_desc[:200],
        'account': 'AT&T',
        'orderId': f'ORD-ATT-{88200 + wf_idx + 1}',
        'slaDeadline': (created + sla_offset).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'created': created.strftime('%Y-%m-%d'),
        'reporter': 'System (auto)',
        'assignee': assignee,
        'status': 'IN PROGRESS',
        'workflowName': wf_name,
        'workflowDesc': wf_desc,
        'currentStep': current_step_raw,
        'errorSnippet': error_snippet,
        'relatedTickets': [],
        'history': history,
        'steps': steps,
        'assigneeGroup': 'Agents' if is_agent else 'Support Team',
        'observability': observability,
    }
    tasks.append(task)

related_groups = {}
for task in tasks:
    prefix = task['workflowName'][:4].upper()
    if prefix not in related_groups:
        related_groups[prefix] = []
    related_groups[prefix].append(task['id'])

for task in tasks:
    prefix = task['workflowName'][:4].upper()
    siblings = [tid for tid in related_groups[prefix] if tid != task['id']]
    task['relatedTickets'] = siblings[:3]

output = {'tasks': tasks}
out_path = r'C:\CursorData\ObsAgents\src\config\common\sky-tasks-att.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"Generated {len(tasks)} tasks")
print(f"Output: {out_path}")

cond_count = sum(1 for t in tasks if any('conditionParent' in s for s in t['steps']))
print(f"Tasks with conditional branches: {cond_count}")

step_labels = set()
for t in tasks:
    for s in t['steps']:
        step_labels.add(s['label'])
print(f"Unique step labels: {len(step_labels)}")
for sl in sorted(step_labels):
    print(f"  - {sl}")
